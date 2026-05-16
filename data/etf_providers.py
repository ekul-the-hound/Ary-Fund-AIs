"""
data/etf_providers.py
=====================
Issuer-specific adapters for ETF holdings ingestion.

Each provider knows how to turn a raw issuer-hosted file (CSV / XLSX)
into a list of canonical ``RawHolding`` records that the loader then
normalizes and writes through ``market_data.set_etf_holdings()``.

Three production adapters
-------------------------
- :class:`IsharesProvider`  — BlackRock iShares CSV
- :class:`SpdrProvider`     — State Street SPDR XLSX / CSV
- :class:`GenericIssuerProvider` — column-mapped CSV for any issuer with
  a flat tabular file (Vanguard, Invesco, etc.). Configure via
  ``ProviderConfig`` rather than subclassing.

Plus :class:`LocalFileProvider` for ingesting a file the user downloaded
manually — the most reliable path in production because issuer URLs
rotate frequently.

Design rules
------------
* Providers do parsing only. No DB writes, no network choice — that's
  the loader's job.
* Each provider returns ``ParseResult`` (success path or structured
  errors). Raising is reserved for programmer errors (bad config).
* ``RawHolding`` is the wire format between providers and the loader.
  The loader is what turns these into canonical records and into the
  ``set_etf_holdings()`` payload shape.

Adding a new issuer
-------------------
1. If the file is a flat CSV with a header row, just register a new
   ``ProviderConfig`` in the loader's ``ETF_PROVIDER_CONFIG`` map and
   point at ``GenericIssuerProvider``. No code needed.
2. If the file has a quirky banner / metadata block / extension, write
   a new subclass of :class:`HoldingsProvider`. Override ``parse_bytes``;
   reuse the shared helpers from this module.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Wire format
# ---------------------------------------------------------------------------


@dataclass
class RawHolding:
    """One holding row as parsed from an issuer file.

    Everything in this dataclass comes straight from the file — no ticker
    resolution, no weight normalization beyond percent→decimal. The loader
    upgrades these into ``CanonicalHolding`` records.
    """
    raw_ticker: Optional[str]          # As printed in the issuer file
    raw_identifier: Optional[str]      # CUSIP / ISIN / SEDOL, whichever the file gives
    identifier_type: Optional[str]     # "cusip" | "isin" | "sedol" | None
    name: Optional[str]
    weight: Optional[float]            # Decimal (0.0725), not percent
    shares: Optional[float]
    market_value: Optional[float]
    extra: dict = field(default_factory=dict)  # Preserved for downstream use


@dataclass
class ParseResult:
    """Outcome of one parse attempt."""
    etf_ticker: str
    issuer: str
    as_of_date: Optional[str]          # ISO YYYY-MM-DD if found
    source_url: str                    # Where this file came from (or "file://...")
    source_type: str                   # "csv" | "xlsx" | "json"
    holdings: list[RawHolding] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors and len(self.holdings) > 0


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


_PCT_RE = re.compile(r"[,\s%]")
_NUM_RE = re.compile(r"[,\s\$]")


def _to_float(raw, allow_pct: bool = False) -> Optional[float]:
    """Best-effort numeric coercion. Returns None on failure.

    Handles common issuer-file noise: thousands separators, currency
    symbols, trailing '%', parentheses for negatives.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in {"-", "—", "N/A", "NA", "n/a", ".", "Cash"}:
        return None
    # Parentheses → negative (accountant style)
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    cleaner = _PCT_RE if allow_pct else _NUM_RE
    s = cleaner.sub("", s)
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _normalize_weight(raw, percent_input: bool = True) -> Optional[float]:
    """Convert an issuer 'Weight' cell to a decimal (0.0725 for 7.25%).

    ``percent_input=True`` (the default) divides by 100. Set False for
    files that already publish decimals (rare).
    """
    v = _to_float(raw, allow_pct=True)
    if v is None:
        return None
    return v / 100.0 if percent_input else v


_DATE_PATTERNS = [
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%b %d, %Y",      # "Apr 30, 2024"
    "%d-%b-%Y",       # "30-Apr-2024"
    "%d %b %Y",
    "%B %d, %Y",      # "April 30, 2024"
    "%Y%m%d",
]


def _parse_date(raw) -> Optional[str]:
    """Try a battery of issuer date formats. Returns ISO YYYY-MM-DD or None."""
    if raw is None:
        return None
    s = str(raw).strip().strip('"').strip()
    if not s:
        return None
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _classify_identifier(raw_id: Optional[str]) -> Optional[str]:
    """Guess the identifier scheme for a raw security id."""
    if not raw_id:
        return None
    s = str(raw_id).strip().upper()
    if not s:
        return None
    if len(s) == 9 and s[:8].replace("X", "").isalnum():
        return "cusip"      # 9 chars, last is check digit
    if len(s) == 12 and s[:2].isalpha():
        return "isin"       # ISO 6166: 2 country letters + 9 alnum + check
    if len(s) == 7 and s.isalnum():
        return "sedol"      # 7 alphanumeric
    return None


def _sniff_csv(text: str) -> str:
    """Detect CSV delimiter. Falls back to comma."""
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _decode_bytes(raw: bytes) -> str:
    """Decode bytes with the encodings issuers actually use."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _find_header_row(rows: list[list[str]], must_contain: set[str]) -> int:
    """Locate the row index that looks like the column header.

    The first row containing all the strings in ``must_contain``
    (case-insensitive, partial match) wins. Returns -1 if not found.
    Issuers like iShares ship 9-10 banner rows before the real header.
    """
    must_lower = {s.lower() for s in must_contain}
    for i, row in enumerate(rows):
        cells = {(c or "").strip().lower() for c in row}
        joined = " ".join(cells)
        if all(any(needle in cell for cell in cells) or needle in joined
               for needle in must_lower):
            return i
    return -1


def _scan_for_date(rows: list[list[str]], max_rows: int = 20) -> Optional[str]:
    """Look through the first ``max_rows`` for a recognizable date.

    Issuers put dates in different forms in the banner: a labeled cell
    ("Fund Holdings as of:,Apr 30, 2024"), a bare ISO date, a single
    cell with both label and value ("As of: 5/15/2024"), or in a "Date:"
    cell. We try each cell directly, then try stripping common label
    prefixes, then try a regex sweep for date-shaped substrings.
    """
    label_strip = re.compile(
        r"^(?:fund\s+holdings\s+)?(?:as\s+of|date|valuation\s+date|"
        r"reporting\s+date)\s*:?\s*",
        re.IGNORECASE,
    )
    # Date-shape regex covers numeric (5/15/2024, 2024-05-15) and
    # month-name (Apr 30, 2024 / 30-Apr-2024) forms.
    date_substr = re.compile(
        r"\d{4}-\d{2}-\d{2}"
        r"|\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
        r"|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4}"
        r"|\d{1,2}[-\s](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-\s]\d{4}",
        re.IGNORECASE,
    )
    for row in rows[:max_rows]:
        for cell in row:
            if not cell:
                continue
            # 1. Direct parse
            d = _parse_date(cell)
            if d:
                return d
            # 2. Strip label prefix and retry
            stripped = label_strip.sub("", str(cell)).strip()
            if stripped and stripped != cell:
                d = _parse_date(stripped)
                if d:
                    return d
            # 3. Regex-find a date-shaped substring anywhere in the cell
            m = date_substr.search(str(cell))
            if m:
                d = _parse_date(m.group(0))
                if d:
                    return d
    return None


def _read_xlsx_first_sheet(raw: bytes) -> list[list[str]]:
    """Read an XLSX from bytes into a list-of-rows.

    Requires openpyxl. Raises ImportError with an actionable message if
    not installed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to parse XLSX holdings files. "
            "Install with: pip install openpyxl"
        ) from e
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[list[str]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in r])
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Base class
# ---------------------------------------------------------------------------


class HoldingsProvider:
    """Abstract provider. Subclasses implement ``build_url`` and ``parse_bytes``."""

    issuer: str = "generic"
    source_type: str = "csv"      # "csv" | "xlsx" | "json"
    source_tag: str = "issuer_csv"  # Passed as `source=` to set_etf_holdings()

    def build_url(self, etf_ticker: str) -> str:
        raise NotImplementedError

    def parse_bytes(
        self, raw: bytes, etf_ticker: str, source_url: str,
    ) -> ParseResult:
        raise NotImplementedError


# ---------------------------------------------------------------------------
# iShares
# ---------------------------------------------------------------------------


class IsharesProvider(HoldingsProvider):
    """BlackRock iShares holdings CSV.

    Real-world structure:

        Fund Holdings as of: ,"Apr 30, 2024",,,,...
        Inception Date: ,"Dec 18, 2009",,,,...
        ...several more banner rows...
        ,,,,,,,,...
        Ticker,Name,Sector,Asset Class,Market Value,Weight (%),Shares,CUSIP,ISIN,...
        AAPL,APPLE INC,Information Technology,Equity,"$58,234,567,890.00",6.85,...
        ...
        The Trust has Cash and/or Derivatives...

    The URL pattern requires a ``product_id`` and human-readable ``slug``.
    BlackRock periodically rotates these — keep them in the loader config
    so updating one ETF doesn't touch parser code.
    """

    issuer = "ishares"
    source_type = "csv"
    source_tag = "ishares_csv"

    URL_TEMPLATE = (
        "https://www.ishares.com/us/products/{product_id}/{slug}/"
        "1467271812596.ajax?fileType=csv&fileName={ticker}_holdings"
        "&dataType=fund"
    )

    HEADER_HINTS = {"ticker", "weight"}

    def __init__(self, product_id: str, slug: str) -> None:
        self.product_id = product_id
        self.slug = slug

    def build_url(self, etf_ticker: str) -> str:
        return self.URL_TEMPLATE.format(
            product_id=self.product_id,
            slug=self.slug,
            ticker=etf_ticker.upper(),
        )

    def parse_bytes(
        self, raw: bytes, etf_ticker: str, source_url: str,
    ) -> ParseResult:
        result = ParseResult(
            etf_ticker=etf_ticker.upper(),
            issuer=self.issuer,
            as_of_date=None,
            source_url=source_url,
            source_type=self.source_type,
        )
        text = _decode_bytes(raw)
        delim = _sniff_csv(text)
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        if not rows:
            result.errors.append("file is empty")
            return result

        result.as_of_date = _scan_for_date(rows)

        header_idx = _find_header_row(rows, self.HEADER_HINTS)
        if header_idx < 0:
            result.errors.append(
                f"column header row not found (looked for {self.HEADER_HINTS}); "
                "issuer may have changed file layout"
            )
            return result

        header = [(c or "").strip() for c in rows[header_idx]]
        col_index = _build_column_index(
            header,
            mapping={
                "ticker":       ["ticker"],
                "name":         ["name"],
                "weight":       ["weight (%)", "weight%", "weight"],
                "shares":       ["shares", "quantity"],
                "market_value": ["market value", "notional value", "value"],
                "cusip":        ["cusip"],
                "isin":         ["isin"],
                "sedol":        ["sedol"],
                "asset_class":  ["asset class", "asset_class"],
                "sector":       ["sector"],
            },
        )
        if "ticker" not in col_index and "cusip" not in col_index:
            result.errors.append(
                "no ticker or CUSIP column in header — refusing to parse"
            )
            return result

        for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any((c or "").strip() for c in row):
                continue  # blank line
            # iShares emits a "The Trust has Cash and/or Derivatives" footer
            joined = " ".join(row).lower()
            if "the trust" in joined or "©" in joined:
                break
            holding = _row_to_holding(row, col_index)
            if holding is None:
                continue
            result.holdings.append(holding)

        if not result.holdings:
            result.errors.append(
                f"parsed 0 holdings from {len(rows) - header_idx - 1} data rows"
            )
        return result


# ---------------------------------------------------------------------------
# SPDR / State Street
# ---------------------------------------------------------------------------


class SpdrProvider(HoldingsProvider):
    """SPDR (State Street) holdings file — XLSX is the canonical format,
    though some ETFs also publish a CSV variant.

    Real-world XLSX structure:

        State Street Global Advisors
        (blank)
        Fund Name:,Technology Select Sector SPDR® Fund
        Date:,4/30/2024
        (blank)
        Name,Ticker,Identifier,SEDOL,Weight,Sector,Shares Held,Local Currency
        APPLE INC,AAPL,037833100,2046251,21.85,Information Technology,...

    The "Identifier" column is a CUSIP for US equities.
    """

    issuer = "spdr"
    source_type = "xlsx"
    source_tag = "spdr_xlsx"

    URL_TEMPLATE = (
        "https://www.ssga.com/us/en/intermediary/library-content/products/"
        "fund-data/etfs/us/holdings-daily-us-en-{ticker_lower}.xlsx"
    )

    HEADER_HINTS = {"name", "ticker", "weight"}

    def build_url(self, etf_ticker: str) -> str:
        return self.URL_TEMPLATE.format(ticker_lower=etf_ticker.lower())

    def parse_bytes(
        self, raw: bytes, etf_ticker: str, source_url: str,
    ) -> ParseResult:
        result = ParseResult(
            etf_ticker=etf_ticker.upper(),
            issuer=self.issuer,
            as_of_date=None,
            source_url=source_url,
            source_type=self.source_type,
        )

        # SPDR sometimes serves CSVs at the same URL — sniff first.
        if raw[:2] == b"PK":
            try:
                rows = _read_xlsx_first_sheet(raw)
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"failed to read XLSX: {e}")
                return result
        else:
            text = _decode_bytes(raw)
            rows = list(csv.reader(io.StringIO(text), delimiter=_sniff_csv(text)))
            result.source_type = "csv"

        if not rows:
            result.errors.append("file is empty")
            return result

        result.as_of_date = _scan_for_date(rows)
        header_idx = _find_header_row(rows, self.HEADER_HINTS)
        if header_idx < 0:
            result.errors.append("column header row not found")
            return result

        header = [(c or "").strip() for c in rows[header_idx]]
        col_index = _build_column_index(
            header,
            mapping={
                "ticker":       ["ticker"],
                "name":         ["name"],
                "weight":       ["weight", "weight (%)"],
                "shares":       ["shares held", "shares", "quantity"],
                "market_value": ["market value", "value"],
                "cusip":        ["identifier", "cusip"],  # SPDR uses 'Identifier'
                "sedol":        ["sedol"],
                "isin":         ["isin"],
                "sector":       ["sector"],
            },
        )
        if "ticker" not in col_index and "cusip" not in col_index:
            result.errors.append(
                "no ticker or identifier column in header"
            )
            return result

        for row in rows[header_idx + 1:]:
            if not any((c or "").strip() for c in row):
                continue
            joined = " ".join(row).lower()
            if joined.startswith("the fund") or "©" in joined:
                break
            holding = _row_to_holding(row, col_index)
            if holding is None:
                continue
            result.holdings.append(holding)

        if not result.holdings:
            result.errors.append("parsed 0 holdings")
        return result


# ---------------------------------------------------------------------------
# Generic / configurable
# ---------------------------------------------------------------------------


@dataclass
class GenericIssuerConfig:
    """Drop-in config for an arbitrary issuer with a flat CSV file.

    Use this when you don't want to subclass — register a new ETF by
    pointing at this provider with a config that names the columns.
    """
    issuer: str                          # "vanguard", "invesco", "wisdomtree", ...
    url_template: str                    # Must contain ``{ticker}`` or ``{ticker_lower}``
    column_map: dict[str, list[str]]     # Canonical name -> list of header aliases
    source_type: str = "csv"
    source_tag: str = "generic_csv"
    header_hints: set[str] = field(default_factory=lambda: {"ticker", "weight"})
    weight_is_percent: bool = True


class GenericIssuerProvider(HoldingsProvider):
    """Config-driven provider for any flat tabular issuer file.

    Build once with a :class:`GenericIssuerConfig`, then it behaves like
    any other provider. Bypasses the need to write a subclass for each
    new issuer.
    """

    source_type = "csv"

    def __init__(self, config: GenericIssuerConfig) -> None:
        self.config = config
        self.issuer = config.issuer
        self.source_type = config.source_type
        self.source_tag = config.source_tag

    def build_url(self, etf_ticker: str) -> str:
        return self.config.url_template.format(
            ticker=etf_ticker.upper(),
            ticker_lower=etf_ticker.lower(),
        )

    def parse_bytes(
        self, raw: bytes, etf_ticker: str, source_url: str,
    ) -> ParseResult:
        result = ParseResult(
            etf_ticker=etf_ticker.upper(),
            issuer=self.issuer,
            as_of_date=None,
            source_url=source_url,
            source_type=self.config.source_type,
        )
        if self.config.source_type == "xlsx" or raw[:2] == b"PK":
            try:
                rows = _read_xlsx_first_sheet(raw)
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"failed to read XLSX: {e}")
                return result
        else:
            text = _decode_bytes(raw)
            rows = list(csv.reader(io.StringIO(text), delimiter=_sniff_csv(text)))
        if not rows:
            result.errors.append("file is empty")
            return result

        result.as_of_date = _scan_for_date(rows)
        header_idx = _find_header_row(rows, self.config.header_hints)
        if header_idx < 0:
            result.errors.append(
                f"header row matching {self.config.header_hints} not found"
            )
            return result

        header = [(c or "").strip() for c in rows[header_idx]]
        col_index = _build_column_index(header, mapping=self.config.column_map)
        if "ticker" not in col_index and "cusip" not in col_index:
            result.errors.append("no ticker or CUSIP column resolved")
            return result

        for row in rows[header_idx + 1:]:
            if not any((c or "").strip() for c in row):
                continue
            h = _row_to_holding(
                row, col_index,
                weight_is_percent=self.config.weight_is_percent,
            )
            if h is None:
                continue
            result.holdings.append(h)

        if not result.holdings:
            result.errors.append("parsed 0 holdings")
        return result


# ---------------------------------------------------------------------------
# Local file
# ---------------------------------------------------------------------------


class LocalFileProvider(HoldingsProvider):
    """Reads a file the user already downloaded. The most reliable
    production path — issuer URLs rotate, manual downloads don't.

    Delegates to the underlying provider for parsing; only the source
    metadata changes (URL becomes ``file://...``).
    """

    def __init__(self, inner: HoldingsProvider) -> None:
        self.inner = inner
        self.issuer = inner.issuer
        self.source_type = inner.source_type
        self.source_tag = inner.source_tag

    def build_url(self, etf_ticker: str) -> str:
        raise NotImplementedError(
            "LocalFileProvider has no URL — pass bytes directly to "
            "the loader's single-file ingestion path"
        )

    def parse_bytes(
        self, raw: bytes, etf_ticker: str, source_url: str,
    ) -> ParseResult:
        return self.inner.parse_bytes(raw, etf_ticker, source_url)


# ---------------------------------------------------------------------------
# Internal: per-row helpers shared by all providers
# ---------------------------------------------------------------------------


def _build_column_index(
    header: list[str],
    mapping: dict[str, list[str]],
) -> dict[str, int]:
    """Map canonical column name → index in the header row.

    ``mapping`` is canonical_name -> [list, of, accepted, header, names]
    (case-insensitive, substring match). First match wins.
    """
    lower = [h.lower().strip() for h in header]
    out: dict[str, int] = {}
    for canonical, aliases in mapping.items():
        for alias in aliases:
            alias_l = alias.lower()
            for i, h in enumerate(lower):
                if h == alias_l or (alias_l in h and len(h) - len(alias_l) <= 4):
                    out[canonical] = i
                    break
            if canonical in out:
                break
    return out


def _row_to_holding(
    row: list[str],
    col_index: dict[str, int],
    weight_is_percent: bool = True,
) -> Optional[RawHolding]:
    """Build a ``RawHolding`` from a parsed CSV row using the resolved
    column index. Returns None if the row has no usable identifier.
    """
    def get(name: str) -> Optional[str]:
        idx = col_index.get(name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        v = v.strip()
        return v if v else None

    raw_ticker = get("ticker")
    cusip = get("cusip")
    isin = get("isin")
    sedol = get("sedol")
    name = get("name")
    weight = _normalize_weight(get("weight"), percent_input=weight_is_percent)
    shares = _to_float(get("shares"))
    mv = _to_float(get("market_value"))

    # Pick the primary identifier and tag its type
    identifier, id_type = None, None
    if cusip:
        identifier, id_type = cusip, "cusip"
    elif isin:
        identifier, id_type = isin, "isin"
    elif sedol:
        identifier, id_type = sedol, "sedol"
    elif raw_ticker:
        identifier, id_type = raw_ticker, "ticker"
    else:
        # No identifier at all — likely a banner or footer row leaked through
        return None

    # Preserve unmapped fields the user might want later
    extra: dict = {}
    for k, v in zip(("sector", "asset_class"), (get("sector"), get("asset_class"))):
        if v:
            extra[k] = v

    return RawHolding(
        raw_ticker=raw_ticker,
        raw_identifier=identifier,
        identifier_type=id_type,
        name=name,
        weight=weight,
        shares=shares,
        market_value=mv,
        extra=extra,
    )
